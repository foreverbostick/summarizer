import sys
import openpyxl as xl
import datetime

# Use sys argument to get path for new PTD/YTD report
path = sys.argv[1]
#path = input('\n** ** What file do you want to open? (Use full path): ** **\n>> ')
# Use 2nd sys argument to select the date
argDate = sys.argv[2]
#argDate = input('\n** ** What day of the month is the file for? (Don\'t include the month): ** **\n>> ')
editDate = argDate
# Path for the revenue summary spreadsheet
rs = sys.argv[3]
#rs = input('\n** ** What file do you want to send the information to? (Use full path): ** **\n>> ')


# Load the PTD/YTD into the program
ptdytd = xl.load_workbook(path)
ptdSheet = ptdytd.active

# Identify the cells with the CC Allowances
visa = ptdSheet['F36'].value
mastercard = ptdSheet['F35'].value
discover = ptdSheet['F34'].value
amex = ptdSheet['F33'].value

#
# Load the RS spreadsheet and make the CC Summary the active workbook
#
ccsWb = xl.load_workbook(rs)
ccsSheet = ccsWb['Credit Card Summary']
rsSheet = ccsWb['Revenue Summary']

#
# Add 4 to the date variable, since the data starts on row 5
#
editDate = int(editDate) + 4

#
# Row corresponding to the entered date
#

#
# Set variables for pasting locations
#
def ccsPaste():
    visaPaste = ccsSheet.cell(row = editDate, column = 3)
    visaPaste.value = visa
    mastercardPaste = ccsSheet.cell(row = editDate, column = 4)
    mastercardPaste.value = mastercard
    discoverPaste = ccsSheet.cell(row = editDate, column = 5)
    discoverPaste.value = discover
    amexPaste = ccsSheet.cell(row = editDate, column = 10)
    amexPaste.value = amex



#####################################
#                                   #
# Revenue summary workbook section  #
#                                   #
#####################################

#
# Variables for revenue summary
#
revenue = ptdSheet['F181'].value
tax = ptdSheet['F192'].value
misc = ptdSheet['F121'].value
outlet = ptdSheet['F130'].value
phone = ptdSheet['F157'].value
cash = ptdSheet['F86'].value
overshort = ptdSheet['F139'].value

#
# Append RS info into RS spreadsheet
#
def rsPaste():
    revenuePaste = rsSheet.cell(row = editDate, column = 3)
    revenuePaste.value = revenue
    taxPaste = rsSheet.cell(row = editDate, column = 4)
    taxPaste.value = tax
    miscPaste = rsSheet.cell(row = editDate, column = 5)
    miscPaste.value = misc
    outletPaste = rsSheet.cell(row = editDate, column = 6)
    outletPaste.value = outlet
    phonePaste = rsSheet.cell(row = editDate, column = 7)
    phonePaste.value = phone
    cashPaste = rsSheet.cell(row = editDate, column = 8)
    cashPaste.value = cash
    overshortPaste = rsSheet.cell(row = editDate, column = 10)
    overshortPaste.value = overshort


#
# See if it's the end of the month, and print a little reminder if so
#
def eom():
    months = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    thisMonth = datetime.date.today().month
    eom = days[thisMonth - 1]
    if eom == int(argDate):
        print('\nIt looks like it\'s the end of the month. Don\'t forget to clean up your files!\n')


#
# Send the info to the RS spreadsheet
#
def wbSave():
    ccsWb.save(filename=rs)

#
# Allow the user to cancel the script in case of an error
#
def cancel():
    choice = input('\n** ** Would you like to continue? Y/N ** **\n>> ')
    if choice.lower == 'y' or 'yes':
        return choice
    else:
        sys.exit(0)

#
# A little message that prints out to let you know the script is complete.
# Will also let you know if there may have been an error with the date.
def complete():
    if str(argDate) == str(rsSheet.cell(row = editDate, column = 2).value) and str(ccsSheet.cell(row = editDate, column = 2).value):
        print('Done!')
    else:
        print('Something didn\'t add up, better check things out manually.')

def doubleCheck():
    print('\n ** ** Before we continue, here are the files you\'re working with: ** **\n')
    print(f'PTDYTD File: {path}\n')
    print(f'Destination File: {rs}\n')
    print(f'\nDate: {argDate}\n')
    print(f'\nVisa: {visa}\n')
    print(f'Mastercard: {mastercard}\n')
    print(f'Discover: {discover}\n')
    print(f'Amex: {amex}\n\n')
    print(f'Revenue: {revenue}\n')
    print(f'Tax: {tax}\n')
    print(f'Misc: {misc}\n')
    print(f'Phone: {phone}\n')
    print(f'Cash: {cash}\n')
    print(f'Over/Short: {overshort}\n')


if __name__ == '__main__':
    ccsPaste()
    rsPaste()
    doubleCheck()
    wbSave()
    eom()
    complete()
